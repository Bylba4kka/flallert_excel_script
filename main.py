import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import json
import os

FILENAME = "file.xlsx"


def replace_words(replacements: dict, phrases: list):
    # print("Заменяем слова..")
    # Новый список для хранения результата
    result = []

    # Итерируем по списку фраз
    for phrase in phrases:
        replaced = False  # Флаг, чтобы проверить, была ли замена
        # print(f"Обрабатываем фразу: {phrase}")
        # Проверяем каждый ключ из словаря
        for key, values in replacements.items():
            if key in phrase:
                # Добавляем оригинальную фразу в результат
                result.append(phrase.replace("(", "").replace(")", ""))
                
                # Если ключ найден в фразе, делаем замены и добавляем в результат
                for value in values:
                    # print("Фраза под замену", phrase)
                    replased_phrase = phrase.replace(key, str(value).strip())
                    # print(f"Заменяем '{phrase}' на '{replased_phrase}'")
                    result.append(replased_phrase)
                
                replaced = True
        
        if not replaced:
            # Если замена не производилась, добавляем исходную фразу
            # print("Исходная фраза", phrase)
            result.append(phrase)

    # Выводим результат
    return result


# Открываем файл
print("Открываем файл..")
workbook = openpyxl.load_workbook(FILENAME, data_only=True)
sheet_script = workbook["script"]
sheet_words = workbook["words"]
sheet_variables = workbook["variables"]
print("Читаем переменные..")
# Читаем переменные
variables_dict = {}
for row in sheet_variables.iter_rows(values_only=True):
    # print(F'Записываем в словарь "{row[0].value}" из листа variables')
    var_name, *var_values = row
    var_values = [var.strip() for var in var_values if var]
    if var_values:
        variables_dict[var_name] = var_values
# print("variables_dict:", variables_dict)
print("Чтение и обработка данных из листа script..")
# Чтение и обработка данных из листа script
for row in sheet_script.iter_rows(min_row=1, max_col=2):
    if row[0].value:
        # print(f"Чтение ячеек из листа script. Значение {row[0].value}. Ячейка: {row}")
        column_a = row[0].value.lower().strip() if row[0].value else None
        column_a = column_a.split("-")

        selected_columns =  []
        for col in column_a:
            try:
                сol_index = column_index_from_string(col)
                selected_columns.append(сol_index)
            except Exception as ex:
                pass
        result_array = []
        for col in selected_columns:
            # print(f"Работаем со столбоцом {get_column_letter(col)} в листе words")
            retry = 1
            multiplier = sheet_words.cell(row=3, column=col).value
            try:
                multiplier = int(multiplier)
            except:
                multiplier = multiplier
            if not isinstance(multiplier, int) | isinstance(multiplier, float):
                print(f"В столбце {get_column_letter(col)} в третьей строке на листе words не цифра, пропускаем столбец. Значение в ячейке: {multiplier}")
                raise Exception
            for cell in sheet_words.iter_rows(min_row=4, min_col=col, max_col=col, max_row=sheet_words.max_row):
                word = cell[0].value
                if word:
                    word = word
                    # print("Работаем над словом:", word, "Ячейка:", cell)
                    if ")-(" in word:
                        retry = len(word.split(")-("))
                    # Умножаем слово на значение из второй строки
                    words_multiplied = [word] * abs(multiplier)
                        
                    for word_variant in words_multiplied:
                        result_array.append(word_variant)
            
        # Замена переменных
        processed_array = result_array
        for i in range(retry):
            processed_array = replace_words(variables_dict, processed_array)
        # print("processed_array", processed_array)
        json_array = []
        for word in processed_array:
            table = str.maketrans("", "", "()")
            word = word.translate(table)
            if "-" in word:
                table = str.maketrans("-", " ")
                json_array.append(word)
                json_array.append(word.translate(table))
                table = str.maketrans("", "", "-")
                json_array.append(word.translate(table))
            else:
                json_array.append(word)
        # print("json_array", json_array)


        # Записываем результат в колонку B
        result = json.dumps(json_array, ensure_ascii=False)
        # print(f"Длина строки {len(result)}")
        filename = f"{row[0].value.replace("-", "")[:250]}.json"
        path = "results/"

        try:
            with open(path + filename, "w", encoding="utf-8") as f:
                f.write(result)
        except FileNotFoundError:
            os.mkdir(path)
            with open(path + filename, "w", encoding="utf-8") as f:
                f.write(result)


        print(f"Сохраняем файл {filename}")

# Сохраняем изменения
print(f"Файлы успешно сохранены..")
